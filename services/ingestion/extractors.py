"""Per-format text extractors. PDF (PyMuPDF), DOCX (python-docx), XLSX (openpyxl
→ markdown tables, all sheets concatenated)."""

from __future__ import annotations

from pathlib import Path


def extract_pdf(path: Path) -> str:
    import fitz  # pymupdf

    with fitz.open(path) as doc:
        parts: list[str] = []
        for page in doc:
            parts.append(page.get_text())
    return "\n".join(parts).strip()


def extract_docx(path: Path) -> str:
    from docx import Document  # python-docx

    doc = Document(str(path))
    paras = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n".join(paras).strip()


def extract_xlsx(path: Path) -> str:
    """Render all sheets as concatenated markdown tables. One string out."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    parts: list[str] = [f"# Workbook: {path.name}", ""]
    for sheet_name in wb.sheetnames:
        rendered = _render_sheet_markdown(wb[sheet_name], sheet_name)
        if rendered:
            parts.append(rendered)
            parts.append("")
    wb.close()
    return "\n".join(parts).strip()


def _render_sheet_markdown(ws, sheet_name: str, *, max_rows: int = 500) -> str:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    rows = _trim_empty_edges(rows)
    if not rows:
        return ""

    truncated = 0
    if len(rows) > max_rows:
        truncated = len(rows) - max_rows
        rows = rows[:max_rows]

    width = max(len(r) for r in rows)
    header = "| " + " | ".join(f"col_{i+1}" for i in range(width)) + " |"
    sep = "| " + " | ".join("---" for _ in range(width)) + " |"
    body = [
        "| " + " | ".join(_fmt_cell(r[i] if i < len(r) else None) for i in range(width)) + " |"
        for r in rows
    ]
    out = [f"## Sheet: {sheet_name}", "", header, sep, *body]
    if truncated:
        out.extend(["", f"_… {truncated} more rows truncated_"])
    return "\n".join(out)


def _trim_empty_edges(rows: list[list]) -> list[list]:
    while rows and all(c is None or c == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    last_nonempty_col = -1
    for i in range(width):
        if any(i < len(r) and r[i] not in (None, "") for r in rows):
            last_nonempty_col = i
    if last_nonempty_col < 0:
        return []
    return [r[: last_nonempty_col + 1] for r in rows]


def _fmt_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if abs(v) >= 1e9:
            return f"{v:.2e}"
        s = f"{v:.4f}".rstrip("0").rstrip(".")
        return s or "0"
    return str(v).replace("|", "\\|").replace("\n", " ")


def extract(path: Path) -> str:
    """Dispatch by suffix. Raises ValueError on unsupported types."""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf(path)
    if suffix == ".docx":
        return extract_docx(path)
    if suffix == ".xlsx":
        return extract_xlsx(path)
    raise ValueError(f"Unsupported file type for extraction: {suffix} ({path.name})")
